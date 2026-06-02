"""Read the A25 Finance Report.xlsx and emit v2/app/a25_seed.json — a static
seed of A25 student rows for the dashboard.

Usage:
    python -m scripts.build_a25_seed \
        "/path/to/A25 Finance Report.xlsx" v2/app/a25_seed.json

The JSON is consumed by /admin/import-a25 on the dashboard. Re-running this
script regenerates the seed; committing the JSON makes it shippable.
"""
from __future__ import annotations
import json, re, sys
from datetime import datetime
from pathlib import Path

try:
    import openpyxl
except ImportError:
    sys.exit("openpyxl is required — `pip install openpyxl` in your venv first")

# Per-location PT cohorts — each tab is a cohort of PT-stream students at
# the named location. "Belfast  Online" (sic — trailing double space in the
# real workbook) is the online cohort.
PT_COHORT_SHEETS = [
    "Swords", "Tallaght", "Cork", "Galway", "Limerick", "Wexford",
    "Belfast  Online",
]
PILATES_SHEET = "Pilates"
DEFERRAL_SHEET = "A25 Deferals "  # sic — trailing space in the real workbook

# Payment-method override sheets — manual transactions that reallocate the
# default-method bucket on the dashboard.
METHOD_OVERRIDE_SHEETS = {
    "CASH PAYMENTS": "Cash",
    "Revolut":       "Revolut",
}

# Column lookup — headers vary slightly between sheet types. We resolve by
# matching the header text, stripping surrounding whitespace.
def _col_map(ws) -> dict[str, int]:
    return {(c.value or "").strip(): i for i, c in enumerate(ws[1]) if c.value}


def _val(row, col_map, key, default=""):
    idx = col_map.get(key)
    if idx is None:
        return default
    v = row[idx]
    return v if v is not None else default


def _parse_date_any(v) -> str:
    """Excel cells can hold either a datetime or a string. Return ISO yyyy-mm-dd
    or empty string."""
    if v is None or v == "":
        return ""
    if isinstance(v, datetime):
        return v.date().isoformat()
    s = str(v).strip()
    for fmt in ("%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y", "%m/%d/%Y", "%d-%-m-%Y"):
        try:
            return datetime.strptime(s, fmt).date().isoformat()
        except ValueError:
            continue
    # Try numeric-only flex (single-digit day or month)
    m = re.match(r"(\d{1,2})[-/](\d{1,2})[-/](\d{4})", s)
    if m:
        d, mo, y = m.groups()
        try:
            return datetime(int(y), int(mo), int(d)).date().isoformat()
        except ValueError:
            return ""
    return ""


def _parse_amount(v) -> float:
    """Lenient float parse. 'Deferral' / blank / strings → 0.0."""
    if v in (None, "", "Deferral"):
        return 0.0
    try:
        return float(v)
    except (TypeError, ValueError):
        s = str(v).replace(",", "").replace("€", "").strip()
        try:
            return float(s)
        except ValueError:
            return 0.0


def _stream_from_qual(qual: str, default: str = "PT") -> str:
    q = (qual or "").lower()
    if "pilates" in q or "reformer" in q:
        return "Pilates"
    return default


def _safe_id(v) -> str:
    """Contact IDs in the workbook are usually floats. Render them as digits-only."""
    if v in (None, ""):
        return ""
    s = str(v).strip()
    # Floats like '37906.0' → '37906'
    if s.endswith(".0"):
        s = s[:-2]
    return s


def _safe_phone(v) -> str:
    if v in (None, ""):
        return ""
    s = str(v).strip()
    if s.endswith(".0"):
        s = s[:-2]
    # Most are stored without the "+" prefix
    if s.isdigit():
        return f"+{s}"
    return s


def _safe_name(v) -> tuple[str, str]:
    """Best-effort first/last split from the Name column."""
    s = (v or "").strip()
    if not s:
        return "", ""
    # Strip a trailing email parenthetical: "Name (name@example.com)"
    s = re.sub(r"\s*\([^)]*@[^)]*\)\s*$", "", s).strip()
    parts = s.split()
    if len(parts) == 1:
        return parts[0], ""
    return parts[0], " ".join(parts[1:])


def _process_cohort_sheet(ws, default_stream: str, is_deferral: bool = False) -> list[dict]:
    cmap = _col_map(ws)
    out = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        contact_id = _safe_id(_val(row, cmap, "Contact ID"))
        if not contact_id:
            continue
        name = _val(row, cmap, "Name", "")
        first, last = _safe_name(name)
        email = (_val(row, cmap, "Email", "") or "").strip()
        phone = _safe_phone(_val(row, cmap, "SMS Number"))
        # Sheet shape varies — Pilates has Pilates-specific column names
        qual = _val(row, cmap, " Qualifications") or _val(row, cmap, "Qualifications") or ""
        if not qual and "Pilates Course Location" in cmap:
            qual = "Pilates Course (EQF Level 4)"
        location = (_val(row, cmap, " Location") or _val(row, cmap, "Pilates Course Location") or "").strip()
        timetable = (_val(row, cmap, " Timetable") or "").strip()
        start_iso = _parse_date_any(
            _val(row, cmap, " Start Date") or _val(row, cmap, "Pilates Course Start Date")
        )
        price = _parse_amount(_val(row, cmap, "Fees Due"))
        spent = _parse_amount(_val(row, cmap, "Paid"))
        plan  = (_val(row, cmap, " Payment Plan") or _val(row, cmap, "Pilates Course Payment Plan") or "").strip()
        method = (_val(row, cmap, "Payment Method") or "").strip()
        stream = _stream_from_qual(str(qual), default=default_stream)
        out.append({
            "contact_id":   contact_id,
            "first_name":   first,
            "last_name":    last,
            "email":        email,
            "phone":        phone,
            "stream":       stream,
            "qualification": str(qual).strip(),
            "location":     location,
            "timetable":    timetable,
            "start_date":   start_iso,
            "price":        price,
            "spent":        spent,
            "payment_plan": plan,
            "payment_method": method,
            "is_deferral":  1 if is_deferral else 0,
        })
    return out


def _process_method_overrides(ws, method: str) -> list[dict]:
    """Sheets like CASH / Revolut list per-contact amounts. Emit one manual
    transaction per contact tagged to that method."""
    cmap = _col_map(ws)
    out = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        contact_id = _safe_id(_val(row, cmap, "Contact ID"))
        if not contact_id:
            continue
        amount = _parse_amount(_val(row, cmap, "Paid"))
        if amount <= 0:
            continue
        d = _parse_date_any(_val(row, cmap, " Start Date") or "")
        out.append({
            "contact_id": contact_id,
            "amount":     amount,
            "method":     method,
            "date":       d or "2025-09-01",
            "note":       f"Imported from A25 Finance Report · {method}",
        })
    return out


def build(xlsx_path: Path) -> dict:
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    students: list[dict] = []
    for s in PT_COHORT_SHEETS:
        if s not in wb.sheetnames:
            print(f"  ! sheet missing: {s!r}", file=sys.stderr)
            continue
        rows = _process_cohort_sheet(wb[s], default_stream="PT")
        students.extend(rows)
        print(f"  {s}: {len(rows)} students", file=sys.stderr)
    rows = _process_cohort_sheet(wb[PILATES_SHEET], default_stream="Pilates")
    students.extend(rows)
    print(f"  {PILATES_SHEET}: {len(rows)} students", file=sys.stderr)
    if DEFERRAL_SHEET in wb.sheetnames:
        rows = _process_cohort_sheet(wb[DEFERRAL_SHEET], default_stream="PT", is_deferral=True)
        students.extend(rows)
        print(f"  {DEFERRAL_SHEET}: {len(rows)} deferrals", file=sys.stderr)

    # Dedupe — a contact can appear in multiple sheets (e.g. an attended cohort
    # plus a deferral list). Keep the highest-paid row per (contact_id, stream).
    seen: dict[tuple[str, str], dict] = {}
    for r in students:
        key = (r["contact_id"], r["stream"])
        prev = seen.get(key)
        if prev is None or r["spent"] > prev["spent"]:
            seen[key] = r
    students = sorted(seen.values(), key=lambda r: (r["stream"], r["location"], r["start_date"], r["last_name"]))

    method_overrides: list[dict] = []
    for sheet_name, method in METHOD_OVERRIDE_SHEETS.items():
        if sheet_name not in wb.sheetnames:
            continue
        rows = _process_method_overrides(wb[sheet_name], method)
        method_overrides.extend(rows)
        print(f"  {sheet_name} → method={method}: {len(rows)} entries", file=sys.stderr)

    return {
        "period":   "A25",
        "students": students,
        "method_overrides": method_overrides,
        "stats": {
            "student_count": len(students),
            "method_override_count": len(method_overrides),
        },
    }


def main() -> int:
    if len(sys.argv) < 3:
        print(__doc__, file=sys.stderr); return 2
    src = Path(sys.argv[1])
    dst = Path(sys.argv[2])
    if not src.exists():
        print(f"source file not found: {src}", file=sys.stderr); return 1
    seed = build(src)
    dst.parent.mkdir(parents=True, exist_ok=True)
    dst.write_text(json.dumps(seed, indent=2, default=str))
    print(f"\n✓ wrote {dst}", file=sys.stderr)
    print(f"  students: {len(seed['students'])}", file=sys.stderr)
    print(f"  method overrides: {len(seed['method_overrides'])}", file=sys.stderr)
    return 0


if __name__ == "__main__":
    sys.exit(main())
